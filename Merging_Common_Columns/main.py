import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Function to merge DataFrames using common columns
def merge_dataframes(dfs):
    # Convert all column headers to lowercase
    for df in dfs:
        df.columns = df.columns.str.lower()

    # Find the common columns in all DataFrames
    common_columns = set(dfs[0].columns)
    for df in dfs[1:]:
        common_columns.intersection_update(df.columns)

    # Check if all common columns are present in all DataFrames
    for df in dfs:
        missing_columns = common_columns - set(df.columns)
        if missing_columns:
            raise KeyError(f"Columns {missing_columns} are not present in DataFrame")

    # Merge the DataFrames using the common columns
    merged_df = pd.concat([df[list(common_columns)] for df in dfs], axis=0)
    return merged_df, list(common_columns)


file_paths = ["/home/husnain/Desktop/A.xlsx", "/home/husnain/Desktop/B.xlsx", "/home/husnain/Desktop/C.xlsx",
              "/home/husnain/Desktop/D.xlsx"]

# Load each Excel file into a pandas DataFrame
dfs = [pd.read_excel(file_path) for file_path in file_paths]

# Merge the DataFrames using common columns
merged_df, common_columns = merge_dataframes(dfs)

# Create a new Workbook object from openpyxl
wb = Workbook()

# Remove the default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Create a new sheet and set its title
sheet_name = "MergedData"  # Replace with your desired sheet name
worksheet = wb.create_sheet(title=sheet_name)

# Convert the DataFrame to a list of lists for writing to the sheet
data = [common_columns] + merged_df.values.tolist()

# Write data to the worksheet
for row in data:
    worksheet.append(row)

# Adjust column width to fit the content
for column in worksheet.columns:
    max_length = 0
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the workbook
output_file_path = "merged_output.xlsx"
wb.save(output_file_path)

print("Merged data saved to:", output_file_path)
